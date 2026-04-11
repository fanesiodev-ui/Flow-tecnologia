"""
Script para criar um balancete de exemplo em Excel.

Gera um arquivo 'balancete_exemplo.xlsx' com dados fictícios
de uma empresa no formato que o sistema consegue ler.

Como usar:
    cd samples
    python criar_balancete.py
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando dependência openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def criar_balancete():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balancete Março 2024"

    # ── Estilos ────────────────────────────────────────────────────────
    azul_escuro = "1E3A5F"
    azul_claro = "D6E4F0"
    verde = "D5F5E3"
    amarelo = "FEF9E7"
    cinza = "F2F3F4"

    fonte_titulo = Font(bold=True, size=14, color="FFFFFF")
    fonte_cabecalho = Font(bold=True, size=10, color="FFFFFF")
    fonte_grupo = Font(bold=True, size=10, color="1E3A5F")
    fonte_normal = Font(size=9)
    fonte_total = Font(bold=True, size=10)

    fill_titulo = PatternFill("solid", fgColor=azul_escuro)
    fill_cabecalho = PatternFill("solid", fgColor="2E86C1")
    fill_receita = PatternFill("solid", fgColor=verde)
    fill_imposto = PatternFill("solid", fgColor="FAD7A0")
    fill_custo = PatternFill("solid", fgColor=azul_claro)
    fill_despesa = PatternFill("solid", fgColor=amarelo)
    fill_total = PatternFill("solid", fgColor=cinza)
    fill_lucro = PatternFill("solid", fgColor="A9DFBF")

    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    centro = Alignment(horizontal="center", vertical="center")
    direita = Alignment(horizontal="right", vertical="center")
    esquerda = Alignment(horizontal="left", vertical="center")

    # ── Dimensões das colunas ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # Código
    ws.column_dimensions["B"].width = 42   # Descrição
    ws.column_dimensions["C"].width = 18   # Saldo R$

    # ── Título ────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = "BALANCETE DE VERIFICAÇÃO — MARÇO/2024"
    ws["A1"].font = fonte_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    ws["A2"] = "Empresa Modelo LTDA  |  CNPJ: 00.000.000/0001-00"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws["A2"].alignment = centro
    ws.row_dimensions[2].height = 18

    # ── Cabeçalho ─────────────────────────────────────────────────────
    cabecalhos = ["Código", "Descrição da Conta", "Saldo (R$)"]
    for col, texto in enumerate(cabecalhos, 1):
        c = ws.cell(row=4, column=col, value=texto)
        c.font = fonte_cabecalho
        c.fill = fill_cabecalho
        c.alignment = centro
        c.border = borda
    ws.row_dimensions[4].height = 20

    # ── Dados do balancete ────────────────────────────────────────────
    # Formato: (código, descrição, valor)
    contas = [
        # RECEITAS
        (None, "── RECEITAS OPERACIONAIS ──", None),
        ("3.1.1", "Receita de Vendas de Produtos", 95_000.00),
        ("3.1.2", "Receita de Prestação de Serviços", 45_000.00),
        ("3.1.3", "Receita Financeira", 10_000.00),
        (None, "TOTAL RECEITAS", 150_000.00),

        # IMPOSTOS
        (None, "", None),
        (None, "── IMPOSTOS E TRIBUTOS ──", None),
        ("4.1.1", "Simples Nacional", 10_500.00),
        ("4.1.2", "ISS — Imposto Sobre Serviços", 2_250.00),
        ("4.1.3", "IRPJ/CSLL Estimativa", 4_500.00),
        ("4.1.4", "PIS e COFINS", 750.00),
        (None, "TOTAL IMPOSTOS", 18_000.00),

        # CUSTOS OPERACIONAIS
        (None, "", None),
        (None, "── CUSTOS OPERACIONAIS ──", None),
        ("5.1.1", "Custo das Mercadorias Vendidas (CMV)", 28_000.00),
        ("5.1.2", "Materiais e Insumos de Produção", 12_000.00),
        ("5.1.3", "Embalagens e Logística", 5_000.00),
        (None, "TOTAL CUSTOS OPERACIONAIS", 45_000.00),

        # DESPESAS ADMINISTRATIVAS
        (None, "", None),
        (None, "── DESPESAS ADMINISTRATIVAS ──", None),
        ("6.1.1", "Salários e Pró-labore", 15_000.00),
        ("6.1.2", "Aluguel do Estabelecimento", 4_500.00),
        ("6.1.3", "Energia Elétrica e Água", 1_800.00),
        ("6.1.4", "Internet e Telefone", 600.00),
        ("6.1.5", "Honorários Contábeis", 1_200.00),
        ("6.1.6", "Softwares e Sistemas", 900.00),
        ("6.1.7", "Outros Gastos Administrativos", 1_000.00),
        (None, "TOTAL DESPESAS ADMINISTRATIVAS", 25_000.00),

        # OUTRAS DESPESAS
        (None, "", None),
        (None, "── OUTRAS DESPESAS ──", None),
        ("7.1.1", "Marketing e Publicidade", 5_500.00),
        ("7.1.2", "Manutenção e Reparos", 3_500.00),
        ("7.1.3", "Despesas com Veículos", 2_000.00),
        ("7.1.4", "Despesas Diversas", 1_000.00),
        (None, "TOTAL OUTRAS DESPESAS", 12_000.00),

        # RESULTADO
        (None, "", None),
        (None, "LUCRO LÍQUIDO DO PERÍODO", 50_000.00),
    ]

    linha = 5
    for codigo, descricao, saldo in contas:
        ws.row_dimensions[linha].height = 17

        # Célula código
        c_cod = ws.cell(row=linha, column=1, value=codigo or "")
        c_cod.font = fonte_normal
        c_cod.alignment = centro
        c_cod.border = borda

        # Célula descrição
        c_desc = ws.cell(row=linha, column=2, value=descricao)
        c_desc.border = borda
        c_desc.alignment = esquerda

        # Célula valor
        c_val = ws.cell(row=linha, column=3, value=saldo)
        c_val.border = borda
        c_val.alignment = direita
        if saldo is not None:
            c_val.number_format = 'R$ #,##0.00'

        # Estiliza cada tipo de linha
        if descricao and "──" in descricao:
            # Linha de grupo/categoria
            c_desc.font = fonte_grupo
            c_desc.fill = PatternFill("solid", fgColor="EAF2FF")
            c_cod.fill = PatternFill("solid", fgColor="EAF2FF")
            c_val.fill = PatternFill("solid", fgColor="EAF2FF")
        elif descricao and "TOTAL" in descricao and "LUCRO" not in descricao:
            # Linha de subtotal
            c_desc.font = fonte_total
            c_val.font = fonte_total
            c_desc.fill = fill_total
            c_cod.fill = fill_total
            c_val.fill = fill_total
        elif descricao and "LUCRO" in descricao:
            # Linha de lucro líquido
            c_desc.font = Font(bold=True, size=10, color="1A5276")
            c_val.font = Font(bold=True, size=10, color="1A5276")
            c_desc.fill = fill_lucro
            c_cod.fill = fill_lucro
            c_val.fill = fill_lucro
        elif descricao == "" or descricao is None:
            # Linha em branco
            pass
        else:
            c_desc.font = fonte_normal

        linha += 1

    # ── Salva o arquivo ────────────────────────────────────────────────
    saida = os.path.join(os.path.dirname(__file__), "balancete_exemplo.xlsx")
    wb.save(saida)
    print(f"\n✅ Balancete criado com sucesso!")
    print(f"📄 Arquivo: {saida}")
    print(f"\nDados do balancete:")
    print(f"  • Faturamento:              R$ 150.000,00")
    print(f"  • Impostos:                  R$ 18.000,00  (12,0%)")
    print(f"  • Custos Operacionais:       R$ 45.000,00  (30,0%)")
    print(f"  • Despesas Administrativas:  R$ 25.000,00  (16,7%)")
    print(f"  • Outras Despesas:           R$ 12.000,00   (8,0%)")
    print(f"  • Lucro Líquido:             R$ 50.000,00  (33,3%)")
    print(f"\nAgora você pode fazer upload deste arquivo no sistema!")


if __name__ == "__main__":
    criar_balancete()
